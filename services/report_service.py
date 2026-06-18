from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from models import Task, Material, SalaryLog
from datetime import datetime, timedelta
from typing import Optional
import logging
import os

logger = logging.getLogger(__name__)

# Папка для отчётов
REPORTS_DIR = "reports"


class ReportService:
    """Сервис для генерации отчётов"""

    @staticmethod
    def generate_daily_report(session: AsyncSession, date: Optional[datetime] = None) -> dict:
        """
        Генерация ежедневного отчёта
        Returns: dict с путями к файлам
        """
        try:
            from fpdf import FPDF
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import matplotlib.pyplot as plt
            import io
            import base64

            if not date:
                date = datetime.now()

            # Создаём папку
            os.makedirs(REPORTS_DIR, exist_ok=True)

            # Получаем данные
            tasks = ReportService._get_tasks_for_date(session, date)
            materials = ReportService._get_materials(session)
            salary = ReportService._get_salary(session, date)

            # Генерируем файлы
            pdf_path = ReportService._generate_pdf(tasks, materials, salary, date)
            excel_path = ReportService._generate_excel(tasks, materials, salary, date)
            charts = ReportService._generate_charts(tasks, materials, salary)

            logger.info(f"Report generated for {date.strftime('%Y-%m-%d')}")

            return {
                "pdf": pdf_path,
                "excel": excel_path,
                "charts": charts,
                "date": date
            }

        except Exception as e:
            logger.error(f"Failed to generate report: {e}")
            return {}

    @staticmethod
    def _get_tasks_for_date(session: AsyncSession, date: datetime) -> list:
        """Получение задач за дату"""
        from sqlalchemy import select
        start_date = date.replace(hour=0, minute=0, second=0)
        end_date = start_date + timedelta(days=1)

        result = session.execute(
            select(Task).where(
                Task.created_at >= start_date,
                Task.created_at < end_date
            )
        )
        return result.scalars().all()

    @staticmethod
    def _get_materials(session: AsyncSession) -> list:
        """Получение материалов"""
        from sqlalchemy import select
        result = session.execute(select(Material))
        return result.scalars().all()

    @staticmethod
    def _get_salary(session: AsyncSession, date: datetime) -> list:
        """Получение зарплаты за дату"""
        from sqlalchemy import select
        start_date = date.replace(hour=0, minute=0, second=0)
        end_date = start_date + timedelta(days=1)

        result = session.execute(
            select(SalaryLog).where(
                SalaryLog.created_at >= start_date,
                SalaryLog.created_at < end_date
            )
        )
        return result.scalars().all()

    @staticmethod
    def _generate_pdf(tasks: list, materials: list, salary: list, date: datetime) -> str:
        """Генерация PDF отчёта"""
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Заголовок
        pdf.cell(200, 10, txt=f"Отчёт за {date.strftime('%d.%m.%Y')}", ln=True, align='C')
        pdf.ln(10)

        # Задачи
        pdf.set_font("Arial", size=14, style='B')
        pdf.cell(200, 10, txt="Задачи:", ln=True)
        pdf.set_font("Arial", size=10)

        for task in tasks[:50]:  # Ограничение
            pdf.cell(200, 8, txt=f"#{task.id} {task.title} - {task.status}", ln=True)

        pdf.ln(5)

        # Материалы
        pdf.set_font("Arial", size=14, style='B')
        pdf.cell(200, 10, txt="Материалы:", ln=True)
        pdf.set_font("Arial", size=10)

        for material in materials[:30]:
            pdf.cell(200, 8, txt=f"{material.name}: {material.quantity} {material.unit.value}", ln=True)

        # Сохраняем
        filename = f"report_{date.strftime('%Y%m%d')}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        pdf.output(filepath)

        return filepath

    @staticmethod
    def _generate_excel(tasks: list, materials: list, salary: list, date: datetime) -> str:
        """Генерация Excel отчёта"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Лист 1: Задачи
        ws1 = wb.active
        ws1.title = "Задачи"
        ws1.append(["ID", "Название", "Статус", "Стоимость", "Дедлайн"])

        for task in tasks:
            ws1.append([
                task.id,
                task.title,
                task.status,
                task.cost or 0,
                task.deadline.strftime('%d.%m.%Y') if task.deadline else ""
            ])

        # Лист 2: Материалы
        ws2 = wb.create_sheet("Материалы")
        ws2.append(["Название", "Остаток", "Единица", "Критический %"])

        for material in materials:
            ws2.append([
                material.name,
                material.quantity,
                material.unit.value,
                material.critical_percent or 10
            ])

        # Лист 3: Зарплата
        ws3 = wb.create_sheet("Зарплата")
        ws3.append(["Пользователь", "Задача", "Сумма", "Дата"])

        for s in salary:
            ws3.append([
                s.user_id,
                s.task_id,
                s.amount,
                s.created_at.strftime('%d.%m.%Y %H:%M')
            ])

        # Сохраняем
        filename = f"report_{date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)

        return filepath

    @staticmethod
    def _generate_charts(tasks: list, materials: list, salary: list) -> dict:
        """Генерация графиков"""
        import matplotlib.pyplot as plt
        import io
        import base64

        charts = {}

        # График 1: Задачи по статусам (pie)
        fig, ax = plt.subplots(figsize=(6, 4))
        status_counts = {}
        for task in tasks:
            status_counts[task.status] = status_counts.get(task.status, 0) + 1

        if status_counts:
            ax.pie(status_counts.values(), labels=status_counts.keys(), autopct='%1.1f%%')
            ax.set_title('Задачи по статусам')
            charts['tasks_pie'] = ReportService._fig_to_base64(fig)
        plt.close()

        # График 2: Задачи по дням (bar)
        fig, ax = plt.subplots(figsize=(6, 4))
        dates = [task.created_at.strftime('%d.%m') for task in tasks]
        if dates:
            ax.hist(dates, bins=min(len(dates), 10))
            ax.set_title('Задачи по дням')
            ax.set_xlabel('Дата')
            ax.set_ylabel('Количество')
            charts['tasks_bar'] = ReportService._fig_to_base64(fig)
        plt.close()

        # График 3: Зарплата (line)
        fig, ax = plt.subplots(figsize=(6, 4))
        if salary:
            amounts = [s.amount for s in salary]
            ax.plot(amounts, marker='o')
            ax.set_title('Зарплата')
            ax.set_xlabel('Задача')
            ax.set_ylabel('Сумма (руб.)')
            charts['salary_line'] = ReportService._fig_to_base64(fig)
        plt.close()

        # График 4: Материалы (heatmap)
        fig, ax = plt.subplots(figsize=(6, 4))
        if materials:
            names = [m.name[:20] for m in materials[:10]]
            quantities = [m.quantity for m in materials[:10]]
            ax.barh(names, quantities)
            ax.set_title('Остатки материалов')
            ax.set_xlabel('Количество')
            charts['materials_bar'] = ReportService._fig_to_base64(fig)
        plt.close()

        return charts

    @staticmethod
    def _fig_to_base64(fig) -> str:
        """Конвертация графика в base64"""
        import io
        import base64
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        img_str = base64.b64encode(buf.read()).decode()
        return img_str